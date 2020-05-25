# -*- coding: utf-8 -*-
import requests,time,os
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
import openpyxl as excel
import datetime
import wx
import wx.xrc

'''
Company对象为公司对象，可获取系统内某一单的所有信息，新建时需输入一个单号
还有一个很大的问题，是用户界面需要加上最终审核建
一键审核功能也需要加上
但是一键上传功能确实比较鸡肋，倾向于不要添加
返金，中介费需要确切金额，而且和AD并不是同时得到，不能加



        这里是吐槽部分
        
            首先需要明确现在公司还有很多制度其实是不完善的。虽无伤大雅，但是对程序完美编写造成了很多问题。此处记录一下
                
                1.理论上兼职是不能上传返金的，但是和机构合作的是可以上传的
                2.返金人没有的是不能返送的   此功能还未实现
                
            
        

'''

class Company(object):
    # 此处规定需传入参数，只有number
    def __init__(self,number):
        self.number=number
        print(number)

        # 返回顾客id 物件名 管理公司 状态
        client_statues=self.get_customer_id()

        self.customer_id=client_statues[0]
        self.client_statues=client_statues[3]

        if client_statues[3]=='该订单未上传或读取问题':


            #申请金状态
            self.applay_fee_statues='该订单未上传或读取问题'
            # 申请金金额
            self.applay_fee_money='该订单未上传或读取问题'
            # 申请金备注
            self.applay_fee_ps='该订单未上传或读取问题'
            # 返金金额
            self.fan_jin='该订单未上传或读取问题'

            # AD金额状态
            self.AD_statue='该订单未上传或读取问题'


            self.money_id = '该订单未上传或读取问题'
            self.ke_shi = '该订单未上传或读取问题'
            self.role = '该订单未上传或读取问题'
            self.check_statue = '该订单未上传或读取问题'
            self.values='该订单未上传或读取问题'



            self.find_number='该订单未上传或读取问题'
            self.statute='该订单未上传或读取问题'
            self.date='该订单未上传或读取问题'
            self.dandang='该订单未上传或读取问题'
            self.wu_jian_ming='该订单未上传或读取问题'
            self.client='该订单未上传或读取问题'
            self.tel='该订单未上传或读取问题'
            self.company='该订单未上传或读取问题'
            self.cai_wu_statues='该订单未上传或读取问题'
        else:
            applay_fee_list = self.get_client_system_values()
            # 申请金状态
            self.applay_fee_statues = applay_fee_list[0]
            # 申请金金额
            self.applay_fee_money = applay_fee_list[1]
            # 申请金备注
            self.applay_fee_ps = applay_fee_list[2]
            # 返金金额
            self.fan_jin = applay_fee_list[3]
            if self.fan_jin == "0":
                self.fan_jin = ""
            # AD金额状态
            self.AD_statue = applay_fee_list[4]
            self.system_agency_fee=applay_fee_list[5]

            money_id_statues = self.get_moey_id()
            self.money_id = money_id_statues[0]
            self.ke_shi = money_id_statues[1]
            self.role = money_id_statues[2]
            self.check_statue = money_id_statues[3]
            self.values = self.get_cai_wu_shu_ju()

            value = self.get_company_value()
            self.find_number = value[0]
            self.statute = value[1]
            self.date = value[2]
            self.dandang = value[3]
            self.wu_jian_ming = value[4]
            self.client = value[5]
            self.tel = value[6]
            self.company = value[8]
            self.cai_wu_statues = self.get_cai_wu_zhuang_tai()

    def get_attr(self):
        attr={"costomer":self.customer_id,"client_statues":self.client_statues,"applay_fee_statues":self.applay_fee_statues,"applay_fee_money":self.applay_fee_money,"applay_fee_ps":self.applay_fee_ps,"fan_jin":self.fan_jin,"values":self.values,"money_id":self.money_id,"ke_shi":self.ke_shi}

    def login_res(self,name,password):
        login_url = 'http://hy.ayqiandu.net/index/login/index.html'
        formData = {'uname': name, 'passwd': password}
        hea = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) PhantomJS/41.0.2272.118 Safari/537.36'}
        s = requests.Session()
        s.post(login_url,data=formData,headers=hea)
        return s

    def get_moey_id(self):
        s = self.login_res(user_number, password)
        html = s.get('http://hy.ayqiandu.net/index/caiwu/index.html?bumen_id=&user_id=&zt=&xm=&wjfh=' + str(
            self.number) + '&cwshsj=&cwshsjend=').text
        soup = BeautifulSoup(html, features="lxml")

        if "LP" in str(self.number):
            tds = soup.find_all(name="td")

            for i in range(len(tds)):

                if tds[i].get_text().strip() == self.number.strip():
                    money_id = tds[i - 1].get_text()
                    # 调整下行的find_next_siblings()[number]中number可以得到对应数据 0为科室，1为社员或兼职，2为名字，3为客户姓名，4为客户状态
                    ke_shi = tds[i + 1].get_text()
                    role = tds[i + 2].get_text()
                    check_statue = tds[i + 5].get_text()

                    return [money_id, ke_shi, role, check_statue]
            return ['该订单未上传', '', "", "该订单未上传"]
        else:

            if soup.find_all(name='td', text=self.number):
                money_id = soup.find_all(name='td', text=self.number)[0].find_previous_sibling().get_text()
                ke_shi = soup.find_all(name='td', text=self.number)[0].find_next_siblings()[0].get_text()
                role = soup.find_all(name='td', text=self.number)[0].find_next_siblings()[1].get_text()
                check_statue = soup.find_all(name='td', text=self.number)[0].find_next_siblings()[4].get_text()
                return [money_id, ke_shi, role, check_statue]
            else:
                return ["该订单未上传", "", "", ""]

        # 想获取更多客户管理系统就在这里更改吧
        #

    def get_cai_wu_shu_ju(self):
        money_id=self.money_id
        s=self.login_res(user_number,password)
        def get_values(money_id,s):
            value=[]
            html = s.get('http://hy.ayqiandu.net/index/caiwu/info/id/'+str(money_id)+'.html').text
            soup = BeautifulSoup(html, features="lxml")
            tbodys=soup.find_all('tbody')
            #下头这俩大兄弟是入金和出金的容纳表格
            # 值得注意的是使用find寻找到的元素依然可以用find进行继续查找

            tds_in1=tbodys[1]
            tds_in2=tbodys[2]

            # 每一行有12个td，整个列表里最后一个为统计的金额，请勿读取

            td1s=tds_in1.find_all('td')
            td2s=tds_in2.find_all('td')


            for i in range(len(td1s)):
                if (i-1)%13==0:
                    mid=[]
                    money=td1s[i].get_text().strip()
                    type=td1s[i+1].get_text().strip()
                    in_out=td1s[i+2].get_text().strip()
                    qing_qiu_shu=td1s[i+3].get_text().strip()
                    qi_yue_shu=td1s[i+4].get_text().strip()
                    time=td1s[i+9].get_text().strip()
                    check_person=td1s[i+10].get_text().strip()
                    id=td1s[i-1].get_text().strip()
                    # 0为类型（AD 其他 中介费 等等） 1为出入金 2为金额 3为请求书 4为契约书 5为审核时间 6为审核人 7为id
                    mid.append(type)
                    mid.append(in_out)
                    mid.append(money)
                    mid.append(qing_qiu_shu)
                    mid.append(qi_yue_shu)
                    mid.append(time)
                    mid.append(check_person)
                    mid.append(id)
                    value.append(mid)

            for i in range(len(td2s)):
                if (i-1)%13==0:
                    mid=[]
                    money=td2s[i].get_text().strip()
                    type=td2s[i+1].get_text().strip()
                    in_out=td2s[i+2].get_text().strip()
                    qing_qiu_shu=td2s[i+3].get_text().strip()
                    qi_yue_shu=td2s[i+4].get_text().strip()
                    time=td2s[i+9].get_text().strip()
                    check_person=td2s[i+10].get_text().strip()
                    id=td2s[i-1].get_text().strip()
                    # 0为类型（AD 其他 中介费 等等） 1为出入金 2为金额 3为请求书 4为契约书 5为审核时间 6为审核人7为id
                    mid.append(type)
                    mid.append(in_out)
                    mid.append(money)
                    mid.append(qing_qiu_shu)
                    mid.append(qi_yue_shu)
                    mid.append(time)
                    mid.append(check_person)
                    mid.append(id)
                    value.append(mid)
            return value
        return get_values(money_id,s)

    def get_company_value(self):
        money_id=self.get_moey_id()[0]
        print(money_id)
        s=self.login_res(user_number,password)
        def get_values(money_id,s):
            value=[]
            html = s.get('http://hy.ayqiandu.net/index/caiwu/info/id/'+str(money_id)+'.html').text
            soup = BeautifulSoup(html, features="lxml")
            tbodys=soup.find_all('tbody')
            #下头这俩大兄弟是入金和出金的容纳表格
            # 值得注意的是使用find寻找到的元素依然可以用find进行继续查找
            tds_in=tbodys[0]


            # 每一行有12个td，整个列表里最后一个为统计的金额，请勿读取
            tds=tds_in.find_all('td')

            for i in range(len(tds)):
                value.append(tds[i].get_text().strip())



            return value
        return get_values(money_id,s)

    def AD_upload(self,money,date,driver,name):


        s=self.login_res('0188','0123456789')
        money_id=self.money_id


        driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(money_id)+".html")
        # 下面两部进入添加界面
        tds=driver.find_elements_by_tag_name("td")
        if tds[0].text=="190107":
            pass
        else:
            cai_wu_sheng_qing=driver.find_element_by_link_text("添加财务申请")
            cai_wu_sheng_qing.click()
            time.sleep(1)
            # 开始添加数据

            # 添加金额
            je=driver.find_element_by_name("je")
            je.send_keys(money)

            # 选择种类
            kind=driver.find_element_by_name('type')
            kind_lei=Select(kind)
            kind_lei.select_by_index(1)

            # 选择类别
            lei_bie=driver.find_element_by_name('qqslx')
            lei_bie_lei=Select(lei_bie)
            lei_bie_lei.select_by_index(1)

            # 上传文件
            upload=driver.find_element_by_name('qqs')


            upload.send_keys('C:\\Users\\user\\Desktop\\upload\\pdf\\'+name+"\\pdf\\总表.pdf")
                    # 日期
            date_value=driver.find_element_by_class_name('date').find_element_by_tag_name('input')
            date_value.click()



            date_value.send_keys(str(date))

            date_value.send_keys(Keys.ENTER)

            nei_rong=driver.find_element_by_name('qqnr')
            nei_rong_lei=Select(nei_rong)
            nei_rong_lei.select_by_index(1)


            # 添加
            btn=driver.find_element_by_id("caiwuadd")
            #btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
            #btn=driver.find_elements_by_tag_name('button')

            btn.click()
            time.sleep(5)






    def agency_fee_upload(self,money,date,driver):


        s=self.login_res('0188','0123456789')
        money_id=self.money_id

        if money_id=="该订单未上传":
            pass
        else:

            driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(money_id)+".html")
        # 下面两部进入添加界面
            cai_wu_sheng_qing=driver.find_element_by_link_text("添加财务申请")
            cai_wu_sheng_qing.click()
            time.sleep(1)
        # 开始添加数据

        # 添加金额
            je=driver.find_element_by_name("je")
            je.send_keys(money)

        # 选择种类
            kind=driver.find_element_by_name('type')
            kind_lei=Select(kind)
            kind_lei.select_by_index(1)

        # 选择类别
            lei_bie=driver.find_element_by_name('qqslx')
            lei_bie_lei=Select(lei_bie)
            lei_bie_lei.select_by_index(3)

        # 上传文件
        #upload=driver.find_element_by_name('qqs')


        #upload.send_keys('C:\\Users\\user\\Desktop\\upload\\agency_fee\\'+str(self.number)+".png")
                # 日期
            date_value=driver.find_element_by_class_name('date').find_element_by_tag_name('input')
            date_value.click()



            date_value.send_keys(str(date))

            date_value.send_keys(Keys.ENTER)

            nei_rong=driver.find_element_by_name('qqnr')
            nei_rong_lei=Select(nei_rong)
            nei_rong_lei.select_by_index(0)


        # 添加
            btn=driver.find_element_by_id("caiwuadd")
        #btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
        #btn=driver.find_elements_by_tag_name('button')

            btn.click()
            time.sleep(5)

    def refund_upload(self,money,date,driver):



        s=self.login_res('0188','0123456789')
        money_id=self.money_id



        if money_id=="该订单未上传":
            pass
        elif "LP" in str(self.number):
            pass
        else:
            print(self.role)
            if self.role=="	兼职":
                pass
            else:
                driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(money_id)+".html")
                # 下面两部进入添加界面
                cai_wu_sheng_qing=driver.find_element_by_link_text("添加财务申请")
                cai_wu_sheng_qing.click()
                time.sleep(1)
                # 开始添加数据

                # 添加金额
                je=driver.find_element_by_name("je")
                je.send_keys(money)

                # 选择种类
                kind=driver.find_element_by_name('type')
                kind_lei=Select(kind)
                kind_lei.select_by_index(2)



                        # 日期
                date_value=driver.find_element_by_class_name('date').find_element_by_tag_name('input')
                date_value.click()



                date_value.send_keys(str(date))

                date_value.send_keys(Keys.ENTER)

                nei_rong=driver.find_element_by_name('qqnr')
                nei_rong_lei=Select(nei_rong)
                nei_rong_lei.select_by_index(0)


                # 添加
                btn=driver.find_element_by_id("caiwuadd")
                #btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
                #btn=driver.find_elements_by_tag_name('button')

                btn.click()
                time.sleep(5)

    def affair_upload(self,date,driver):

        s=self.login_res('0188','0123456789')
        money_id=self.money_id

        if money_id=="该订单未上传":
            pass
        else:

            driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(money_id)+".html")
            # 下面两部进入添加界面
            cai_wu_sheng_qing=driver.find_element_by_link_text("添加财务申请")
            cai_wu_sheng_qing.click()
            time.sleep(1)
            # 开始添加数据

            # 添加金额
            je=driver.find_element_by_name("je")
            je.send_keys("10000")

            # 选择种类
            kind=driver.find_element_by_name('type')
            kind_lei=Select(kind)
            kind_lei.select_by_index(2)




                    # 日期
            date_value=driver.find_element_by_class_name('date').find_element_by_tag_name('input')
            date_value.click()



            date_value.send_keys(str(date))

            date_value.send_keys(Keys.ENTER)

            nei_rong=driver.find_element_by_name('qqnr')
            nei_rong_lei=Select(nei_rong)
            nei_rong_lei.select_by_index(2)


            # 添加
            btn=driver.find_element_by_id("caiwuadd")
            #btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
            #btn=driver.find_elements_by_tag_name('button')

            btn.click()
            time.sleep(5)

    def get_cai_wu_zhuang_tai(self):



        zhongjiefei=[]
        xian_jin_AD=[]
        AD=[]
        in_qi_ta=[]
        fan_jin=[]
        out_qi_ta=[]
        if "LP" in str(self.number):
            mid_number=str(self.number)[2:6]
            print(mid_number)
        else:
            mid_number=str(self.number)[0:4]
        if 2001>int(mid_number)>=1907 :
            k=0.3
        else:
            k=0.5
        for i in range(len(self.values)):

            if self.values[i][1]=="入金":
                if self.values[i][0]=="中介费":
                    zhongjiefei.append(self.values[i])

                elif self.values[i][0]=="现金AD":
                    xian_jin_AD.append(self.values[i])

                elif self.values[i][0]=="AD":
                    AD.append(self.values[i])

                elif self.values[i][0]=="其他":
                    in_qi_ta.append(self.values[i])

            elif self.values[i][1]=="出金":
                if self.values[i][0]=="返金":
                    fan_jin.append(self.values[i])

                elif self.values[i][0]=="其他":
                    out_qi_ta.append(self.values[i])

    # mid_value用于存放中间值
        mid_value={}

    # 判断中介费状态
        if len(zhongjiefei)==0:
            if self.system_agency_fee=="0" or "":
                zhong_jie_fei_statues ="无中介费"
                out_money=0
            else:
                zhong_jie_fei_statues='中介费未上传'
                out_money="客户申请系统内金额为"+self.system_agency_fee
            mid_value["中介费"]=[out_money,zhong_jie_fei_statues]
        elif len(zhongjiefei)==1:
            zhong_jie_fei_statues="仅有一笔中介费"
            mid_value["中介费"]=[zhongjiefei[0][2],zhong_jie_fei_statues]
        else:
            zhong_jie_fei_statues="有"+str(len(zhongjiefei))+"笔中介费"
            mid_value["中介费"]=['',zhong_jie_fei_statues]

    # 判断现金AD状态
        if len(xian_jin_AD)==0:
            xian_jin_AD_statues='现金AD未上传'
            mid_value["现金AD"]=['',xian_jin_AD_statues]
        elif len(xian_jin_AD)==1:
            xian_jin_AD_statues="仅有一笔现金AD"
            mid_value["现金AD"]=[xian_jin_AD[0][2],xian_jin_AD_statues]
        else:
            xian_jin_AD_statues="有"+str(len(xian_jin_AD))+"笔现金AD"
            mid_value["现金AD"]=['',xian_jin_AD_statues]

    # 判断AD状态
        if len(AD)==0:
            AD_statues='AD未上传'
            if len(self.AD_statue)==1:
                AD_statues="AD为0"
            mid_value["AD"]=['',AD_statues]
        elif len(AD)==1:
            AD_statues="仅有一笔AD"
            mid_value["AD"]=[AD[0][2],AD_statues]
        else:
            AD_statues="有"+str(len(AD))+"笔AD"
            ad_zong=0
            for i in range(len(AD)):
                ad_zong+=int(AD[i][2])
            mid_value["AD"]=[ad_zong,AD_statues]

    # 判断返金状态
        if len(fan_jin)==0:
            if self.fan_jin =="" or self.fan_jin==0:
                fan_jin_statues='此单无返金'
                fan_jin_out_money=0
            else:
                fan_jin_statues='返金未上传'
                fan_jin_out_money="系统内返金状态为"+self.fan_jin
            mid_value["返金"]=[fan_jin_out_money,fan_jin_statues]

        elif len(fan_jin)==1:
            fan_jin_statues="仅有一笔返金"
            mid_value["返金"]=[fan_jin[0][2],fan_jin_statues]

        else:
            fan_jin_statues="有"+str(len(fan_jin))+"笔返金"
            mid_value["返金"]=['',fan_jin_statues]


    # 判断总务扣款状态
        if len(out_qi_ta)==0:
            out_qi_ta_statues='总务扣款未上传'
            mid_value["总务扣款"]=['',out_qi_ta_statues]
        elif len(out_qi_ta)==1:
            out_qi_ta_statues="仅有一笔总务扣款"
            mid_value["总务扣款"]=[out_qi_ta[0][2],out_qi_ta_statues]
        else:
            out_qi_ta_statues="有"+str(len(out_qi_ta))+"笔总务扣款"
            zong_wu_zong = 0
            for i in range(len(out_qi_ta)):
                zong_wu_zong += int(out_qi_ta[i][2])

            mid_value["总务扣款"]=[zong_wu_zong,out_qi_ta_statues]

    # 判断AD和现金AD状态

        if AD_statues=="仅有一笔AD":

            if xian_jin_AD_statues=='现金AD未上传':
                AD_result=True
                AD_value=AD[0][2]

            else:
                AD_result=False
        elif AD_statues=="AD未上传":
            if xian_jin_AD_statues=='仅有一笔现金AD':
                AD_result=True
                AD_value=xian_jin_AD[0][2]
            else:
                AD_result=False

        else:
            AD_result=False
            AD_value=0

        if AD_result and zhong_jie_fei_statues=="仅有一笔中介费" and fan_jin_statues=="仅有一笔返金":
            try:
                if (int(zhongjiefei[0][2])+int(AD_value)+770)*k>=int(fan_jin[0][2]):
                    mid_value["对比结果"]="返金符合要求"
                else:
                    mid_value["对比结果"]="返金不符合要求"
            except:
                mid_value["对比结果"] = "可能存在多笔AD或同时存在AD和现金AD"
        elif fan_jin_statues=="此单无返金":
            mid_value["对比结果"]="此单无返金"
        elif zhong_jie_fei_statues=="中介费未上传" or "无中介费":
            try:
                fan_jin=fan_jin[0][2]
            except:
                fan_jin=0
            try:
                if (int(AD_value)+770)*k>=int(fan_jin):
                    mid_value["对比结果"]="返金符合要求"
                else:
                    mid_value["对比结果"]="返金不符合要求"
            except:
                mid_value["对比结果"] = "可能存在多笔AD或同时存在AD和现金AD"
        else:
            mid_value["对比结果"]="缺少输入项"


        return mid_value

    def check_angency_fee(self,driver,date):

        if self.cai_wu_statues["中介费"][1]=="仅有一笔中介费":
            for i in range(len(self.values)):
                if self.values[i][0]=="中介费":
                    id=self.values[i][7]
                    check_person=self.values[i][6]
                    driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(self.money_id)+".html")
                    btns=driver.find_elements_by_tag_name("button")
            if check_person == "":
                for i in range(len(btns)):
                    if btns[i].get_attribute("data-target")=="#exampleModal"+str(id):
                        btns[i].click()
                        time.sleep(1)
                        # 点击审核成功
                        driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][1]/label[@class='radio-inline'][1]/input").click()
                        # 输入日期
                        date_btn=driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][3]/div[@class='input-group date form_date col-md-12']/input[@class='form-control']")
                        date_btn.click()
                        date_btn.clear()
                        date_btn.send_keys(date)
                        date_btn.send_keys(Keys.ENTER)

                        # 点击提交
                        driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-footer']/button[@class='btn btn-primary']").click()
                        time.sleep(5)
                        break


    def check_office_fee(self,driver,date):



        if self.cai_wu_statues["总务扣款"][1]=="仅有一笔总务扣款":

            for i in range(len(self.values)):
                if self.values[i][1]=="出金":
                    if self.values[i][0]=="其他":
                        id=self.values[i][7]
                        check_person=self.values[i][6]

                    driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(self.money_id)+".html")
                    btns=driver.find_elements_by_tag_name("button")
            if check_person == "":
                for i in range(len(btns)):

                    if btns[i].get_attribute("data-target")=="#exampleModal"+str(id):
                        btns[i].click()
                        time.sleep(1)
                         #点击审核成功
                        driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][1]/label[@class='radio-inline'][1]/input").click()
                        # 输入日期
                        date_btn=driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][3]/div[@class='input-group date form_date col-md-12']/input[@class='form-control']")
                        date_btn.click()
                        date_btn.clear()
                        date_btn.send_keys(date)
                        date_btn.send_keys(Keys.ENTER)

                        # 点击提交
                        driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-footer']/button[@class='btn btn-primary']").click()
                        time.sleep(5)
                        break
            else:
                pass

    def check_fan_jin(self,driver,date):
        if self.cai_wu_statues["对比结果"]=="返金符合要求":
            for i in range(len(self.values)):
                if self.values[i][0]=="返金":
                    id=self.values[i][7]
                    check_person=self.values[i][6]
                    driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(self.money_id)+".html")
                    btns=driver.find_elements_by_tag_name("button")
            if check_person == "":
                for i in range(len(btns)):
                        if btns[i].get_attribute("data-target")=="#exampleModal"+str(id):
                            btns[i].click()
                            time.sleep(1)
                            # 点击审核成功
                            driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][1]/label[@class='radio-inline'][1]/input").click()
                            # 输入日期
                            date_btn=driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][3]/div[@class='input-group date form_date col-md-12']/input[@class='form-control']")
                            date_btn.click()
                            date_btn.clear()
                            date_btn.send_keys(date)
                            date_btn.send_keys(Keys.ENTER)

                            # 点击提交
                            driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal"+id+"']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-footer']/button[@class='btn btn-primary']").click()
                            time.sleep(5)
                            break

    def check_final(self,driver):
        driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(self.money_id)+".html")
        kind = driver.find_element_by_id('selectkhzt')
        kind_lei = Select(kind)
        kind_lei.select_by_index(1)

        btn = driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/form[@id='form']/table[@class='table table-bordered']/tbody/tr/td/div[@class='form-group'][2]/div[@class='col-sm-offset-1 col-sm-3']/button[@class='btn btn-default']")
        btn.click()
        time.sleep(5)

    def get_customer_id(self):
        s = self.login_res(user_number, password)
        html = s.get("http://hy.ayqiandu.net/index/shenqing/index.html?user_id=&khlb=&wjfh=" + str(
            self.number) + "&wujianming=&xm=&glgsdh=&glgsm=&shengri=").text
        soup = BeautifulSoup(html, features="lxml")

        this_number = str(self.number).strip()

        tds = soup.find_all(name="td")

        for i in range(len(tds)):

            if tds[i].get_text().strip() == this_number:
                customer_id = tds[i - 3].get_text()
                # 调整下行的find_next_siblings()[number]中number可以得到对应数据 0为科室，1为社员或兼职，2为名字，3为客户姓名，4为客户状态
                wu_jian_ming = tds[i + 1].get_text()
                guan_li_gong_si = tds[i + 2].get_text()
                client_statues = tds[i + 6].get_text()
                return [customer_id, wu_jian_ming, guan_li_gong_si, client_statues]
        return ['该订单未上传', '', '', '该订单未上传或读取问题']

    def check_user_statues(self,driver):
        driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/"+str(self.money_id)+".html")
        kind=driver.find_element_by_id('selectkhzt')
        kind_lei=Select(kind)
        kind_lei.select_by_index(1)

        btns=driver.find_elements_by_tag_name("button")
        k=[]
        for i in range(len(btns)):
            if btns[i].text=="审核":
                k.append(btns[i])
        btn=k[len(k)-1]
        btn.click()
        time.sleep(4)


    # 想获取更多客户管理系统就在这里更改吧
    def get_client_system_values(self):

        s = self.login_res(user_number, password)
        html = s.get("http://hy.ayqiandu.net/index/shenqing/info/id/" + str(self.customer_id) + ".html").text
        soup = BeautifulSoup(html, features="lxml")

        apply_money_statues = soup.find_all("select")[4].find_all("option")
        for i in range(len(apply_money_statues)):
            if len(apply_money_statues[i].attrs) == 2:
                apply_statues = apply_money_statues[i].get_text()
        a = soup.find_all("input")

        fan_jin = soup.find_all("input")[27].attrs['value']
        money = soup.find_all("input")[20].attrs['value']
        ps = soup.find_all("input")[22].attrs['value']
        AD_statue = soup.find_all("input")[21].attrs['value']
        angency_fee = soup.find_all("input")[23].attrs['value']
        rent = soup.find_all("input")[34].attrs['value']
        management_fee = soup.find_all("input")[35].attrs['value']
        # 定金状态 定金金额 备注 返金金额 AD金额 中介费 房租 管理费
        return [apply_statues, money, ps, fan_jin, AD_statue, angency_fee, rent, management_fee]


# 增加数据
def add_line(number,path,length,add):
    # 打开excel
    name = os.path.join(path)
    wb = excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
    company_this = Company(number)

    ws.cell(row=length+1+add, column=1, value=number)
    ws.cell(row=length+1+add, column=2, value=company_this.cai_wu_statues["中介费"][1])
    ws.cell(row=length+1+add, column=3, value=company_this.cai_wu_statues["现金AD"][1])
    ws.cell(row=length+1+add, column=4, value=company_this.cai_wu_statues["AD"][1])
    ws.cell(row=length+1+add, column=5, value=company_this.cai_wu_statues["返金"][1])
    ws.cell(row=length+1+add, column=6, value=company_this.cai_wu_statues["总务扣款"][1])
    ws.cell(row=length+1+add, column=7, value=company_this.statute)

    wb.save(path)


# 修改数据
def chang_line(number, path, line, value):
    # 打开excel
    name = os.path.join(path)
    wb = excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])

    ws.cell(row=line + 2, column=1, value=number)
    if "中介费" in value:
        ws.cell(row=line + 2, column=2, value=value["中介费"])
    if "现金AD" in value:
        ws.cell(row=line + 2, column=3, value=value["现金AD"])
    if "AD" in value:
        ws.cell(row=line + 2, column=4, value=value["AD"])
    if "返金" in value:
        ws.cell(row=line + 2, column=5, value=value["返金"])
    if "总务扣款" in value:
        ws.cell(row=line + 2, column=6, value=value["总务扣款"])
    if "状态" in value:
        ws.cell(row=line + 2, column=7, value=value["状态"])

    wb.save(path)

# 读取信息表
def get_database(path):
    # 打开excel
    name = os.path.join(path)
    wb = excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])

def get_database(path,value_list):
    # 打开excel
    name=os.path.join(path)
    wb=excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    print(sheets)
    ws = wb.get_sheet_by_name(sheets[0])
    # 创建数据存储
    content_mid=[]
    # 遍历行获取数据并输出
    rows = ws.rows
    for row in rows:
        now_1 = time.time()

        k=[] # 注意这里 放在第二个循环内是不可以的

        for cell in row:
            k.append(cell.value)


        value_list.append(k)
    value_list.remove(value_list[0])


database=[]
get_database("C:\\Users\\user\\Desktop\\upload\\excel\\单子详情.xlsx",database)





def login_sele(name,password,driver):
    driver.get('http://hy.ayqiandu.net/index/login/index.html')
    user = driver.find_element_by_name('uname')
    password1 = driver.find_element_by_name('passwd')
    user.send_keys(name)
    password1.send_keys(password)
    btn=driver.find_element_by_class_name('loginbtn')
    btn.click()

def get_final_date(start_date):
    start_date=str(start_date)
    year=start_date[0:4]
    if "-" in start_date:
        list=start_date.split("-")
    elif "/" in start_date:
        list=start_date.split("/")

    month=int(list[1])




    if month in [1,3,5,7,8,10,12]:
        day=31
    elif month==2:
        day=28
    else:
        day=30
    return str(year)+"-"+str(month)+"-"+str(day)


value_content=[]

def get_excel_value(path,value_list):
    # 打开excel
    name=os.path.join(path)
    wb=excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    print(sheets)
    ws = wb.get_sheet_by_name(sheets[0])
    # 创建数据存储
    content_mid=[]
    # 遍历行获取数据并输出
    rows = ws.rows
    for row in rows:
        k=[] # 注意这里 放在第二个循环内是不可以的

        for cell in row:
            k.append(cell.value)


        value_list.append(k)

def put_cai_wu_value(path,value_content):
    # 打开excel
    name=os.path.join(path)
    wb=excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
    # 创建数据存储
    content_mid=[]
    # 遍历行获取数据并输出

    # 此处把excel第一行的标题插入
    ws.cell(row=1, column=1, value="单号")
    ws.cell(row=1, column=2, value="担当")
    ws.cell(row=1, column=3, value="科室")
    ws.cell(row=1, column=4, value="中介费金额")
    ws.cell(row=1, column=5, value="中介费状态")
    ws.cell(row=1, column=6, value="中介费审核日期")
    ws.cell(row=1, column=7, value="现金AD金额")
    ws.cell(row=1, column=8, value="现金AD状态")
    ws.cell(row=1, column=9, value="现金AD审核日期")
    ws.cell(row=1, column=10, value="AD金额")
    ws.cell(row=1, column=11, value="AD状态")
    ws.cell(row=1, column=12, value="AD审核日期")
    ws.cell(row=1, column=13, value="返金金额")
    ws.cell(row=1, column=14, value="返金状态")
    ws.cell(row=1, column=15, value="返金审核日期")
    ws.cell(row=1, column=16, value="总务扣款金额")
    ws.cell(row=1, column=17, value="总务扣款状态")
    ws.cell(row=1, column=18, value="总务扣款审核日期")
    ws.cell(row=1, column=19, value="是否可返")


    for i in range(len(value_content)):
        print(i)
        print(value_content[i])
        # 此处更改几个人的科室信息
        if value_content[i-1][2]=="袁泉":
            value_content[i-1][3]="营业3课"
        elif value_content[i-1][2]=="	孫昊東":
            value_content[i-1][3] = "营业3课"
        elif value_content[i-1][2]=="	宋博倫":
            value_content[i-1][3] = "营业3课"
        elif value_content[i-1][2]=="	周正":
            value_content[i-1][3] = "营业5课"

        # 开始循环写入数据

        ws.cell(row = i+2, column = 1,value=value_content[i-1][0])
        ws.cell(row = i+2, column = 2,value=value_content[i-1][2])
        ws.cell(row = i+2, column = 3,value=value_content[i-1][3])

        if value_content[i][1]["中介费"][0]=="":
            value_content[i][1]["中介费"][0]=0
        ws.cell(row = i+2, column = 4,value=value_content[i-1][1]['中介费'][0])
        ws.cell(row = i+2, column = 5,value=value_content[i-1][1]['中介费'][1])
        ws.cell(row = i+2, column = 6,value=value_content[i-1][5])

        if value_content[i-1][1]["现金AD"][0]=="":
            value_content[i-1][1]["现金AD"][0]=0
        elif value_content[i-1][1]["现金AD"][0]==None:
            value_content[i-1][1]["现金AD"][0]=0
        ws.cell(row = i+2, column = 7,value=value_content[i-1][1]['现金AD'][0])
        ws.cell(row = i+2, column = 8,value=value_content[i-1][1]['现金AD'][1])
        ws.cell(row = i+2, column = 9,value=value_content[i-1][8])

        if value_content[i-1][1]["AD"][0]=="":
            value_content[i-1][1]["AD"][0]=0
        ws.cell(row = i+2, column = 10,value=value_content[i-1][1]['AD'][0])
        ws.cell(row = i+2, column = 11,value=value_content[i-1][1]['AD'][1])
        ws.cell(row = i+2, column = 12,value=value_content[i-1][4])

        if value_content[i-1][1]["返金"][0]=="":
            value_content[i-1][1]["返金"][0]=0
        ws.cell(row = i+2, column = 13,value=value_content[i-1][1]['返金'][0])
        ws.cell(row = i+2, column = 14,value=value_content[i-1][1]['返金'][1])
        ws.cell(row = i+2, column = 15,value=value_content[i-1][6])


        if value_content[i-1][1]["总务扣款"][0]=="":
            value_content[i-1][1]["总务扣款"][0]=0
        ws.cell(row = i+2, column = 16,value=value_content[i-1][1]['总务扣款'][0])
        ws.cell(row = i+2, column = 17,value=value_content[i-1][1]['总务扣款'][1])
        ws.cell(row = i+2, column = 18,value=value_content[i-1][7])

        ws.cell(row = i+2, column = 19,value=value_content[i-1][1]["对比结果"])
        ws.cell(row = i+2, column = 21,value=value_content[i-1][9])

        ws.cell(row=i + 2, column=20, value=value_content[i - 1][10])
        ws.cell(row=i + 2, column=22, value=value_content[i - 1][11])

        if value_content[i-1][1]["对比结果"]=="返金不符合要求":
            ft = Font(color=RED)
            ws.cell(row = i+2, column = 19).font=ft

        #value_content.append(k)
    wb.save(path)


# 此处用来最终审核
driver=webdriver.Chrome()
login_sele(user_number,password,driver)
final_check=[] 
get_excel_value("C:\\Users\\user\\Desktop\\input\\AD_list.xlsx",final_check)



for i in range(len(final_check)):
    this_number=Company(final_check[i][0])
    if this_number.check_statue=="财务申请":
        this_number.check_final(driver)
'''

class MyFrame1 ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"登陆", pos = wx.DefaultPosition, size = wx.Size( 195,142 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.DefaultSize, wx.DefaultSize )

        bSizer1 = wx.BoxSizer( wx.VERTICAL )
        

        gbSizer1 = wx.GridBagSizer( 0, 0 )
        gbSizer1.SetFlexibleDirection( wx.BOTH )
        gbSizer1.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )

        self.m_staticText3 = wx.StaticText( self, wx.ID_ANY, u"账号：", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText3.Wrap( -1 )
        gbSizer1.Add( self.m_staticText3, wx.GBPosition( 0, 0 ), wx.GBSpan( 1, 1 ), wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL, 5 )

        self.m_textCtrl3 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        gbSizer1.Add( self.m_textCtrl3, wx.GBPosition( 0, 1 ), wx.GBSpan( 1, 1 ), wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL, 5 )


        bSizer1.Add( gbSizer1, 0, wx.EXPAND, 5 )

        gbSizer2 = wx.GridBagSizer( 0, 0 )
        gbSizer2.SetFlexibleDirection( wx.BOTH )
        gbSizer2.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )

        self.m_staticText4 = wx.StaticText( self, wx.ID_ANY, u"密码：", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText4.Wrap( -1 )
        gbSizer2.Add( self.m_staticText4, wx.GBPosition( 0, 0 ), wx.GBSpan( 1, 1 ), wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL, 5 )

        self.m_textCtrl4 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, wx.TE_PASSWORD )
        gbSizer2.Add( self.m_textCtrl4, wx.GBPosition( 0, 1 ), wx.GBSpan( 1, 1 ), wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL, 5 )


        bSizer1.Add( gbSizer2, 0, wx.EXPAND, 5 )

        self.m_button1 = wx.Button( self, wx.ID_ANY, u"登陆", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer1.Add( self.m_button1, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        self.SetSizer( bSizer1 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.m_button1.Bind( wx.EVT_BUTTON, self.login )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def login( self, event ):

        account=self.m_textCtrl3.GetValue()

        password=self.m_textCtrl4.GetValue()

        if account=="" or password=="":
            self.SetTitle("请输入账号密码")

        else:
            global account_password
            account_password=[account,password]
            frame1.Close()
            frame2.Show(True)

class MyFrame2 ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 314,265 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.DefaultSize, wx.DefaultSize )

        bSizer3 = wx.BoxSizer( wx.VERTICAL )

        gbSizer5 = wx.GridBagSizer( 0, 0 )
        gbSizer5.SetFlexibleDirection( wx.BOTH )
        gbSizer5.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )

        self.m_button8 = wx.Button( self, wx.ID_ANY, u"上传中介费", wx.DefaultPosition, wx.DefaultSize, 0 )
        gbSizer5.Add( self.m_button8, wx.GBPosition( 0, 0 ), wx.GBSpan( 1, 1 ), wx.ALL, 5 )

        self.m_button9 = wx.Button( self, wx.ID_ANY, u"上传返金", wx.DefaultPosition, wx.DefaultSize, 0 )
        gbSizer5.Add( self.m_button9, wx.GBPosition( 0, 1 ), wx.GBSpan( 1, 1 ), wx.ALL, 5 )

        self.m_button13 = wx.Button( self, wx.ID_ANY, u"总务扣款", wx.DefaultPosition, wx.DefaultSize, 0 )
        gbSizer5.Add( self.m_button13, wx.GBPosition( 0, 2 ), wx.GBSpan( 1, 1 ), wx.ALL, 5 )


        bSizer3.Add( gbSizer5, 0, wx.EXPAND, 5 )

        gbSizer6 = wx.GridBagSizer( 0, 0 )
        gbSizer6.SetFlexibleDirection( wx.BOTH )
        gbSizer6.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )

        self.m_button10 = wx.Button( self, wx.ID_ANY, u"审核中介费", wx.DefaultPosition, wx.DefaultSize, 0 )
        gbSizer6.Add( self.m_button10, wx.GBPosition( 0, 0 ), wx.GBSpan( 1, 1 ), wx.ALL, 5 )

        self.m_button11 = wx.Button( self, wx.ID_ANY, u"审核返金", wx.DefaultPosition, wx.DefaultSize, 0 )
        gbSizer6.Add( self.m_button11, wx.GBPosition( 0, 1 ), wx.GBSpan( 1, 1 ), wx.ALL, 5 )

        self.m_button14 = wx.Button( self, wx.ID_ANY, u"审核总务扣款", wx.DefaultPosition, wx.DefaultSize, 0 )
        gbSizer6.Add( self.m_button14, wx.GBPosition( 0, 2 ), wx.GBSpan( 1, 1 ), wx.ALL, 5 )


        bSizer3.Add( gbSizer6, 0, wx.EXPAND, 5 )

        self.m_button12 = wx.Button( self, wx.ID_ANY, u"下载财务信息", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer3.Add( self.m_button12, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.m_textCtrl3 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 310,100 ), wx.TE_MULTILINE)
        bSizer3.Add( self.m_textCtrl3, 0, wx.ALL, 5 )
        self.m_textCtrl3.SetValue("使用说明：\n"
                                  "1.第一行为上传按钮\n"
                                  "2.第二行为审核按钮\n"
                                  "3.第三行为财务信息下载按钮，点击即可实现下载当前系统内指定单号财务信息\n"
                                  "4.点击按钮后程序进入运行状态，请勿再次点击。运行完毕后即可点击")


        self.SetSizer( bSizer3 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.m_button8.Bind( wx.EVT_BUTTON, self.up_agency_fee )
        self.m_button9.Bind( wx.EVT_BUTTON, self.up_fan_jin )
        self.m_button13.Bind( wx.EVT_BUTTON, self.up_office_fee )
        self.m_button10.Bind( wx.EVT_BUTTON, self.check_agency_fee )
        self.m_button11.Bind( wx.EVT_BUTTON, self.check_fan_jin )
        self.m_button14.Bind( wx.EVT_BUTTON, self.check_office_fee )
        self.m_button12.Bind( wx.EVT_BUTTON, self.down_values )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def up_agency_fee( self, event ):
        self.m_textCtrl3.Clear()
        value_list=[]
        # 取得中介费上传单号
        get_excel_value("C:\\Users\\user\\Desktop\\input\\agency_fee.xlsx",value_list)

        date_this=datetime.date.today()
        driver=webdriver.Chrome()
        login_sele(account_password[0],account_password[1],driver)

        for i in range(len(value_list)):

            number=value_list[i][0]

            money=value_list[i][1]

            aim=True
            for j in range(len(database)):
                if str(number)==str(database[j][0]):
                    zhong_jie_fei=database[j][1]
                    xian_jin_AD=database[j][2]
                    AD=database[j][3]
                    fan_jin=database[j][4]
                    zong_wu=database[j][5]
                    zhuang_tai=database[j][6]
                    add=0
                    if zhuang_tai=='财务审核成功':
                        pass
                    elif zhuang_tai=="财务申请":

                        if zhong_jie_fei=="中介费未上传":



                            company_this=Company(number)

                            this_date=str(datetime.date.today())
                            money_this=company_this.cai_wu_statues["中介费"][0]
                            money_statues=company_this.cai_wu_statues["中介费"][1]
                            if company_this.check_statue=="财务审核成功":
                                pass
                            else:
                                if money_statues=="中介费未上传":
                                    print("中介费状态"+money_statues)
                                    company_this.agency_fee_upload(money,this_date,driver)
                                    str_begin=self.m_textCtrl3.GetValue()
                                    str_in=str_begin+"\n"+"物件番号"+str(number)+"中介费已上传"
                                    self.m_textCtrl3.SetValue(str_in)
                                    chang_line(number, "C:\\Users\\user\\Desktop\\upload\\excel\\单子详情.xlsx", j,
                                               {"中介费": "仅有一笔中介费"})


                                elif money_statues=="仅有一笔中介费":
                                    if int(money_this) != int(money):
                                        str_begin=self.m_textCtrl3.GetValue()
                                        str_in=str_begin+"\n"+"物件番号"+str(number)+"中介费金额不符"
                                        self.m_textCtrl3.SetValue(str_in)
                                elif money_statues=="该订单未上传或读取问题":
                                    str_begin = self.m_textCtrl3.GetValue()
                                    str_in = str_begin + "\n" + "物件番号" + str(number) + "该订单未上传或读取问题"
                                    self.m_textCtrl3.SetValue(str_in)
                                else:
                                    str_begin=self.m_textCtrl3.GetValue()
                                    str_in=str_begin+"\n"+"物件番号"+str(number)+"有多笔中介费"
                                    self.m_textCtrl3.SetValue(str_in)
                        elif zhong_jie_fei=="无中介费":
                            print(zhong_jie_fei)



                            company_this=Company(number)

                            this_date=str(datetime.date.today())
                            money_this=company_this.cai_wu_statues["中介费"][0]
                            money_statues=company_this.cai_wu_statues["中介费"][1]
                            if company_this.check_statue=="财务审核成功":
                                pass
                            else:
                                if money_statues=="中介费未上传" or "无中介费":
                                    print("中介费状态"+money_statues)
                                    company_this.agency_fee_upload(money,this_date,driver)
                                    str_begin=self.m_textCtrl3.GetValue()
                                    str_in=str_begin+"\n"+"物件番号"+str(number)+"中介费已上传"
                                    self.m_textCtrl3.SetValue(str_in)
                                    chang_line(number, "C:\\Users\\user\\Desktop\\upload\\excel\\单子详情.xlsx", j,
                                               {"中介费": "仅有一笔中介费"})


                                elif money_statues=="仅有一笔中介费":
                                    if int(money_this) != int(money):
                                        str_begin=self.m_textCtrl3.GetValue()
                                        str_in=str_begin+"\n"+"物件番号"+str(number)+"中介费金额不符"
                                        self.m_textCtrl3.SetValue(str_in)
                                elif money_statues=="该订单未上传或读取问题":
                                    str_begin = self.m_textCtrl3.GetValue()
                                    str_in = str_begin + "\n" + "物件番号" + str(number) + "该订单未上传或读取问题"
                                    self.m_textCtrl3.SetValue(str_in)
                                else:
                                    str_begin=self.m_textCtrl3.GetValue()
                                    str_in=str_begin+"\n"+"物件番号"+str(number)+"有多笔中介费"
                                    self.m_textCtrl3.SetValue(str_in)
                    aim=False
                    break
            if aim:
                print(number)
                company_this = Company(number)
                this_date = str(datetime.date.today())
                money_this = company_this.cai_wu_statues["中介费"][0]
                money_statues = company_this.cai_wu_statues["中介费"][1]
                if company_this.check_statue == "财务审核成功":
                    pass
                else:
                    if money_statues == "中介费未上传":
                        print("中介费状态" + money_statues)
                        company_this.agency_fee_upload(money, this_date, driver)
                        str_begin = self.m_textCtrl3.GetValue()
                        str_in = str_begin + "\n" + "物件番号" + str(number) + "中介费已上传"
                        self.m_textCtrl3.SetValue(str_in)
                        add += 1
                        add_line(number, "C:\\Users\\user\\Desktop\\upload\\excel\\单子详情.xlsx", j, add)
                    elif money_statues == "仅有一笔中介费":
                        if int(money_this) != int(money):
                            str_begin = self.m_textCtrl3.GetValue()
                            str_in = str_begin + "\n" + "物件番号" + str(number) + "中介费金额不符"
                            self.m_textCtrl3.SetValue(str_in)
                    elif money_statues == "该订单未上传或读取问题":
                        str_begin = self.m_textCtrl3.GetValue()
                        str_in = str_begin + "\n" + "物件番号" + str(number) + "该订单未上传或读取问题"
                        self.m_textCtrl3.SetValue(str_in)
                    else:
                        str_begin = self.m_textCtrl3.GetValue()
                        str_in = str_begin + "\n" + "物件番号" + str(number) + "有多笔中介费"
                        self.m_textCtrl3.SetValue(str_in)








    def up_fan_jin( self, event ):
        self.m_textCtrl3.Clear()
        value_list=[]
        get_excel_value("C:\\Users\\user\\Desktop\\input\\fan_jin.xlsx",value_list)
        date_this=datetime.date.today()
        driver=webdriver.Chrome()
        login_sele(account_password[0],account_password[1],driver)
        add=0

        for i in range(len(value_list)):

            number=value_list[i][0]
            money=value_list[i][1]
            aim = True
            for j in range(len(database)):

                if str(number)==str(database[j][0]):
                    zhong_jie_fei=database[j][1]
                    xian_jin_AD=database[j][2]
                    AD=database[j][3]
                    fan_jin=database[j][4]
                    zong_wu=database[j][5]
                    zhuang_tai=database[j][6]
                    if zhuang_tai=='财务审核成功':
                        print()
                        pass
                    elif zhuang_tai=="财务申请":
                        if fan_jin=="返金未上传":

                            company_this=Company(number)
                            this_date=str(datetime.date.today())
                            print(company_this.cai_wu_statues)
                            money_this=company_this.cai_wu_statues["返金"][0]
                            money_statues=company_this.cai_wu_statues["返金"][1]

                            if company_this.check_statue=="财务审核成功":
                                pass
                            else:
                                if money_statues=="返金未上传":
                                    company_this.refund_upload(money,this_date,driver)
                                    str_begin=self.m_textCtrl3.GetValue()
                                    str_in=str_begin+"\n"+"物件番号"+str(number)+"返金已上传"
                                    self.m_textCtrl3.SetValue(str_in)
                                    chang_line(number, "C:\\Users\\user\\Desktop\\upload\\excel\\单子详情.xlsx", j, {"返金": "仅有一笔返金"})
                                    aim = False
                                    break
                                elif money_statues=="仅有一笔返金":
                                    print(company_this.role)
                                    if abs(int(money_this)-int(money))<400:
                                        str_begin=self.m_textCtrl3.GetValue()
                                        str_in=str_begin+"\n"+"物件番号"+str(number)+"返金金额不符"
                                        self.m_textCtrl3.SetValue(str_in)
                                elif money_statues=="该订单未上传或读取问题":
                                    str_begin = self.m_textCtrl3.GetValue()
                                    str_in = str_begin + "\n" + "物件番号" + str(number) + "该订单未上传或读取问题"
                                    self.m_textCtrl3.SetValue(str_in)
                                else:
                                    print(company_this.role)
                                    str_begin=self.m_textCtrl3.GetValue()
                                    str_in=str_begin+"\n"+"物件番号"+str(number)+"有多笔返金"
                                    self.m_textCtrl3.SetValue(str_in)

            if aim:
                company_this = Company(number)
                this_date = str(datetime.date.today())

                money_this = company_this.cai_wu_statues["返金"][0]
                money_statues = company_this.cai_wu_statues["返金"][1]

                if company_this.check_statue == "财务审核成功":
                    pass
                else:
                    if money_statues == "返金未上传":
                        company_this.refund_upload(money, this_date, driver)
                        str_begin = self.m_textCtrl3.GetValue()
                        str_in = str_begin + "\n" + "物件番号" + str(number) + "返金已上传"
                        self.m_textCtrl3.SetValue(str_in)
                        add+=1
                        add_line(number,"C:\\Users\\user\\Desktop\\upload\\excel\\单子详情.xlsx",j,add)
                    elif money_statues == "仅有一笔返金":
                        print(company_this.role)
                        if abs(int(money_this) - int(money)) < 400:
                            str_begin = self.m_textCtrl3.GetValue()
                            str_in = str_begin + "\n" + "物件番号" + str(number) + "返金金额不符"
                            self.m_textCtrl3.SetValue(str_in)
                    elif money_statues == "该订单未上传或读取问题":
                        str_begin = self.m_textCtrl3.GetValue()
                        str_in = str_begin + "\n" + "物件番号" + str(number) + "该订单未上传或读取问题"
                        self.m_textCtrl3.SetValue(str_in)
                    else:
                        print(company_this.role)
                        str_begin = self.m_textCtrl3.GetValue()
                        str_in = str_begin + "\n" + "物件番号" + str(number) + "有多笔返金"
                        self.m_textCtrl3.SetValue(str_in)

    def up_office_fee( self, event ):
        self.m_textCtrl3.Clear()
        value_list=[]
        get_excel_value("C:\\Users\\user\\Desktop\\input\\office_fee.xlsx",value_list)
        date_this=datetime.date.today()
        driver=webdriver.Chrome()
        login_sele(account_password[0],account_password[1],driver)
        for i in range(len(value_list)):

            number=value_list[i][0]
            money=value_list[i][1]
            company_this=Company(number)
            money_this = company_this.cai_wu_statues["总务扣款"][0]
            money_statues = company_this.cai_wu_statues["总务扣款"][1]
            print(i+1)
            print(number)
            print(company_this.cai_wu_statues)

            if money_statues=="总务扣款未上传":
                this_date=str(datetime.date.today())
                company_this.affair_upload(this_date,driver)
                str_begin=self.m_textCtrl3.GetValue()
                str_in=str_begin+"\n"+"物件番号"+str(number)+"总务扣款已上传"
                self.m_textCtrl3.SetValue(str_in)
                chang_line(number, "C:\\Users\\user\\Desktop\\upload\\excel\\单子详情.xlsx",i, {"总务扣款": "仅有一笔总务扣款"})
            elif money_statues=="该订单未上传或读取问题":
                str_begin = self.m_textCtrl3.GetValue()
                str_in = str_begin + "\n" + "物件番号" + str(number) + "该订单未上传或读取问题"
                self.m_textCtrl3.SetValue(str_in)


        print("总务扣款上传完毕")

    def check_agency_fee( self, event ):
        self.m_textCtrl3.Clear()
        value_list=[]
        value_out=[]
        get_excel_value("C:\\Users\\user\\Desktop\\input\\AD_list.xlsx",value_list)
        driver=webdriver.Chrome()
        login_sele(account_password[0],account_password[1],driver)
        for i in range(len(value_list)):
            number=value_list[i][0]
            date_this=str(value_list[i][2])

            date_final=get_final_date(date_this)
            company_this=Company(number)
            company_this.check_angency_fee(driver,date_final)
            str_begin=self.m_textCtrl3.GetValue()
            str_in=str_begin+"\n"+"物件番号"+str(number)+"中介费已审核"
            self.m_textCtrl3.SetValue(str_in)

    def check_fan_jin( self, event ):
        self.m_textCtrl3.Clear()
        value_list=[]
        value_out=[]
        get_excel_value("C:\\Users\\user\\Desktop\\input\\AD_list.xlsx",value_list)
        driver=webdriver.Chrome()
        login_sele(account_password[0],account_password[1],driver)
        for i in range(len(value_list)):
            number=value_list[i][0]
            date_this=str(value_list[i][2])

            date_final=get_final_date(date_this)
            company_this=Company(number)


            company_this.check_fan_jin(driver,date_final)
            str_begin=self.m_textCtrl3.GetValue()
            str_in=str_begin+"\n"+"物件番号"+str(number)+"返金已审核"
            self.m_textCtrl3.SetValue(str_in)


    def check_office_fee( self, event ):
        self.m_textCtrl3.Clear()
        value_list=[]
        value_out=[]
        get_excel_value("C:\\Users\\user\\Desktop\\input\\AD_list.xlsx",value_list)
        driver=webdriver.Chrome()
        login_sele(account_password[0],account_password[1],driver)
        for i in range(len(value_list)):
            number=value_list[i][0]
            date_this=str(value_list[i][2])

            date_final=get_final_date(date_this)
            company_this=Company(number)

            company_this.check_office_fee(driver,date_final)
            str_begin=self.m_textCtrl3.GetValue()
            str_in=str_begin+"\n"+"物件番号"+str(number)+"总务扣款已审核"
            self.m_textCtrl3.SetValue(str_in)


    def down_values( self, event ):
        self.m_textCtrl3.Clear()

        value_list=[]
        value_out=[]
        get_excel_value("C:\\Users\\user\\Desktop\\input\\AD_list.xlsx",value_list)
        print(value_list)
        for i in range(len(value_list)):
            number=value_list[i][0]
            print(number)
            company_this=Company(number)

            values_this=company_this.values
            date=""
            date_agency_fee=""
            date_fan_jin=""
            date_office_fee=""
            date_crash_AD=""
            for i in range(len(values_this)):
                if values_this[i][0]=="AD":
                    date=values_this[i][5]
                    break
            for i in range(len(values_this)):
                if values_this[i][0]=="中介费":
                    date_agency_fee=values_this[i][5]
                    break
            for i in range(len(values_this)):
                if values_this[i][0]=="现金AD":
                    date_crash_AD=values_this[i][5]
                    break
            for i in range(len(values_this)):
                if values_this[i][0]=="返金":
                    date_fan_jin=values_this[i][5]
                    break

            for i in range(len(values_this)):
                if values_this[i][0]=="其他":
                    if values_this[i][1]=="出金":
                        date_office_fee=values_this[i][5]
                        break



            k=[number,company_this.cai_wu_statues,company_this.dandang,company_this.ke_shi,date,date_agency_fee,date_fan_jin,date_office_fee,date_crash_AD,company_this.client,company_this.role,company_this.wu_jian_ming]

            value_out.append(k)

        print(value_out)
        put_cai_wu_value("C:\\Users\\user\\Desktop\\output\\cai_wu_xin_xi.xlsx",value_out)
        self.m_textCtrl3.SetValue("财务信息已下载完成")

    def final(self,event):
        driver=webdriver.Chrome()
        login_sele(user_number,password,driver)
        value_list=[]
        get_excel_value("C:\\Users\\user\\Desktop\\input\\AD_list.xlsx",value_list)

        for i in range(len(value_list)):
            number=value_list[i][0]
            date_this=str(value_list[i][2])

            date_final=get_final_date(date_this)
            company_this=Company(number)
            company_this.check_user_statues(driver)

def final():
    driver=webdriver.Chrome()
    login_sele(user_number,password,driver)
    value_list=[]
    get_excel_value("C:\\Users\\user\\Desktop\\input\\AD_list.xlsx",value_list)

    for i in range(len(value_list)):
        number=value_list[i][0]
        date_this=str(value_list[i][2])

        date_final=get_final_date(date_this)
        company_this=Company(number)
        company_this.check_user_statues(driver)
#final()




app = wx.App(False)
frame1 = MyFrame1(None)
frame2 = MyFrame2(None)

frame1.Show(True)
#start the applications
app.MainLoop()


'''

